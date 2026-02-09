"""Utilities for generating official compliance reports via the lowenergy.jp API.

Uses the official v380 (Ver.3.8) endpoints:
  - reportFromInputSheets  → 公式様式PDF
  - computeFromInputSheets → 公式計算結果JSON
See: https://api.lowenergy.jp/model/1/v380/
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import io
import logging

import openpyxl
import requests

logger = logging.getLogger(__name__)

# ── API settings ────────────────────────────────────────────────────────────
API_BASE = "https://api.lowenergy.jp/model/1/v380"
API_REPORT = f"{API_BASE}/reportFromInputSheets"
API_COMPUTE = f"{API_BASE}/computeFromInputSheets"
EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SMALLMODEL_UPLOAD_UNSUPPORTED_MESSAGE = (
    "小規模版（SMALLMODEL）原本Excelの直接アップロードは未対応です。"
    "公式BEI画面から入力して送信するか、MODEL形式の入力シートをご利用ください。"
)

# Directory that stores the official Excel templates bundled with this repo.
TEMPLATE_DIR = Path(__file__).resolve().parents[2] / "Excel　書式"
SMALL_TEMPLATE = TEMPLATE_DIR / "SMALLMODEL_inputSheet_for_Ver3.8_beta.xlsx"
STANDARD_TEMPLATE = TEMPLATE_DIR / "MODEL_inputSheet_for_Ver3.8_beta.xlsx"


# ── 様式A: 基本情報 — 固定セルマッピング ────────────────────────────────────
# key → (sheet_name, cell_coordinate)
# MODEL版とSMALLMODEL版でシート名が異なる (様式A vs 様式SA)
FORM_A_MAPPING: Dict[str, Tuple[str, str]] = {
    "sheet_date":            ("様式A_基本情報", "C3"),
    "author":                ("様式A_基本情報", "C4"),
    "building_name":         ("様式A_基本情報", "C6"),
    "prefecture":            ("様式A_基本情報", "D7"),
    "city":                  ("様式A_基本情報", "G7"),
    "region":                ("様式A_基本情報", "C9"),
    "solar_region":          ("様式A_基本情報", "C10"),
    "total_area":            ("様式A_基本情報", "C11"),
    "code_symbol":           ("様式A_基本情報", "E13"),
    "code_use":              ("様式A_基本情報", "E14"),
    "building_type":         ("様式A_基本情報", "E15"),
    "room_type":             ("様式A_基本情報", "E16"),
    "calc_floor_area":       ("様式A_基本情報", "C17"),
    "ac_floor_area":         ("様式A_基本情報", "C18"),
    "floors_above":          ("様式A_基本情報", "D19"),
    "floors_below":          ("様式A_基本情報", "G19"),
    "total_height":          ("様式A_基本情報", "C20"),
    "perimeter":             ("様式A_基本情報", "C21"),
    "non_ac_core_direction": ("様式A_基本情報", "D22"),
    "non_ac_core_length":    ("様式A_基本情報", "G22"),
}

# SMALLMODEL版のシート名マッピング
SMALL_SHEET_MAP: Dict[str, str] = {
    "様式A_基本情報": "様式SA_基本情報",
    "様式B1_開口部仕様": "様式SB1_開口部仕様",
    "様式B2_断熱仕様": "様式SB2_断熱仕様",
    "様式B3_外皮": None,  # SMALLMODEL には存在しない
    "様式C1_空調熱源": "様式SC1_空調熱源",
    "様式C2_空調外気処理": "様式SC2_空調外気処理",
    "様式C3_空調ポンプ": None,  # SMALLMODEL には存在しない
    "様式C4_空調送風機": None,  # SMALLMODEL には存在しない
    "様式D_換気": "様式SD_換気",
    "様式E_照明": "様式SE_照明",
    "様式F_給湯": "様式SF_給湯",
    "様式G_昇降機": None,  # SMALLMODEL には存在しない
    "様式H_太陽光発電": "様式SH_太陽光発電",
    "様式I_コージェネレーション設備": None,  # SMALLMODEL には存在しない
}

# ── テーブル形式シートの列定義 (行11〜) ────────────────────────────────────
# key → Excel column letter
TABLE_COLUMNS: Dict[str, Dict[str, str]] = {
    "B1": {
        "sheet": "様式B1_開口部仕様",
        "columns": {
            "name": "A", "width": "B", "height": "C", "area": "D",
            "window_type": "E", "glass_type": "F",
            "glass_u_value": "G", "glass_shgc": "H",
            "window_u_value": "I", "window_shgc": "J",
        },
    },
    "B2": {
        "sheet": "様式B2_断熱仕様",
        "columns": {
            "name": "A", "part_class": "B", "input_method": "C",
            "material_category": "D", "material_detail": "E",
            "conductivity": "F", "thickness": "G", "u_value": "H",
        },
    },
    "B3": {
        "sheet": "様式B3_外皮",
        "columns": {
            "name": "A", "direction": "B", "width": "C", "height": "D",
            "area": "E", "insulation_name": "F", "window_name": "G",
            "window_count": "H", "has_blind": "I",
            "shade_coeff_cooling": "J", "shade_coeff_heating": "K",
        },
    },
    "C1": {
        "sheet": "様式C1_空調熱源",
        "columns": {
            "name": "A", "type": "B", "count": "C",
            "capacity_cooling": "D", "capacity_heating": "E",
            "power_cooling": "F", "power_heating": "G",
            "fuel_cooling": "H", "fuel_heating": "I",
        },
    },
    "C2": {
        "sheet": "様式C2_空調外気処理",
        "columns": {
            "name": "A", "count": "B",
            "supply_airflow": "C", "exhaust_airflow": "D",
            "heat_exchange_eff_cooling": "E", "heat_exchange_eff_heating": "F",
            "auto_bypass": "G", "preheat_stop": "H",
        },
    },
    "C3": {
        "sheet": "様式C3_空調ポンプ",
        "columns": {
            "name": "A", "count": "B", "flow_rate": "C",
            "variable_flow": "D", "min_flow_input": "E", "min_flow_ratio": "F",
        },
    },
    "C4": {
        "sheet": "様式C4_空調送風機",
        "columns": {
            "name": "A", "count": "B", "airflow": "C",
            "variable_airflow": "D", "min_airflow_input": "E", "min_airflow_ratio": "F",
        },
    },
    "D": {
        "sheet": "様式D_換気",
        "columns": {
            "room_name": "A", "room_type": "B", "floor_area": "C",
            "method": "D", "equipment_name": "E", "count": "F",
            "airflow": "G", "motor_power": "H",
            "high_eff_motor": "I", "inverter": "J", "airflow_control": "K",
        },
    },
    "E": {
        "sheet": "様式E_照明",
        "columns": {
            "room_name": "A", "room_type": "B", "floor_area": "C",
            "room_height": "D", "fixture_name": "E",
            "power_per_unit": "F", "count": "G",
            "occupancy_sensor": "H", "daylight_control": "I",
            "schedule_control": "J", "initial_illuminance": "K",
        },
    },
    "F": {
        "sheet": "様式F_給湯",
        "columns": {
            "system_name": "A", "use_type": "B", "source_name": "C",
            "count": "D", "heating_capacity": "E",
            "power_consumption": "F", "fuel_consumption": "G",
            "insulation_level": "H", "water_saving": "I",
        },
    },
    "G": {
        "sheet": "様式G_昇降機",
        "columns": {
            "name": "A", "control_type": "B",
        },
    },
    "H": {
        "sheet": "様式H_太陽光発電",
        "columns": {
            "system_name": "A", "cell_type": "B", "installation_mode": "C",
            "capacity_kw": "D", "panel_direction": "E", "panel_angle": "F",
        },
    },
    "I": {
        "sheet": "様式I_コージェネレーション設備",
        "columns": {
            "name": "A", "rated_output": "B", "count": "C",
            "gen_eff_100": "D", "gen_eff_75": "E", "gen_eff_50": "F",
            "heat_eff_100": "G", "heat_eff_75": "H", "heat_eff_50": "I",
            "heat_recovery_for": "J",
        },
    },
}

TABLE_START_ROW = 11  # 全テーブルシート共通


def _resolve_sheet_name(model_sheet: str, is_small: bool) -> Optional[str]:
    """MODEL版シート名をSMALLMODEL版に変換。存在しない場合はNone。"""
    if not is_small:
        return model_sheet
    return SMALL_SHEET_MAP.get(model_sheet, model_sheet)


def _write_form_a(wb: openpyxl.Workbook, building: Dict[str, Any], is_small: bool) -> None:
    """様式A (基本情報) を固定セルに書き込む。"""
    for key, (model_sheet, cell) in FORM_A_MAPPING.items():
        value = building.get(key)
        if value is None:
            continue
        sheet_name = _resolve_sheet_name(model_sheet, is_small)
        if sheet_name is None or sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found for key '%s'", sheet_name, key)
            continue
        wb[sheet_name][cell] = value
        logger.debug("様式A: %s[%s] = %s", sheet_name, cell, value)


def _write_table_rows(
    wb: openpyxl.Workbook,
    form_key: str,
    rows: List[Dict[str, Any]],
    is_small: bool,
) -> None:
    """テーブル形式シート(B1〜I)にデータ行を展開。"""
    if not rows:
        return

    table_def = TABLE_COLUMNS.get(form_key)
    if table_def is None:
        logger.warning("Unknown table form key: %s", form_key)
        return

    model_sheet = table_def["sheet"]
    sheet_name = _resolve_sheet_name(model_sheet, is_small)

    if sheet_name is None:
        logger.info("Sheet %s skipped (SMALLMODEL does not have it)", model_sheet)
        return
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found in workbook", sheet_name)
        return

    ws = wb[sheet_name]
    columns = table_def["columns"]

    for i, row_data in enumerate(rows):
        excel_row = TABLE_START_ROW + i
        if excel_row > 1010:
            logger.warning("Table %s exceeded max rows (1010)", form_key)
            break
        for field_key, col_letter in columns.items():
            value = row_data.get(field_key)
            if value is None:
                continue
            ws[f"{col_letter}{excel_row}"] = value
        logger.debug("様式%s: row %d written", form_key, excel_row)


def _write_data_to_workbook(workbook: openpyxl.Workbook, input_data: Dict[str, Any]) -> None:
    """input_data dict → Excelワークブックに書き込む。

    input_data の構造:
    {
        "building": { ... 様式A のフィールド ... },
        "windows": [ { ... }, ... ],       # 様式B1
        "insulations": [ { ... }, ... ],   # 様式B2
        "envelopes": [ { ... }, ... ],     # 様式B3
        "heat_sources": [ { ... }, ... ],  # 様式C1
        "outdoor_air": [ { ... }, ... ],   # 様式C2
        "pumps": [ { ... }, ... ],         # 様式C3
        "fans": [ { ... }, ... ],          # 様式C4
        "ventilations": [ { ... }, ... ],  # 様式D
        "lightings": [ { ... }, ... ],     # 様式E
        "hot_waters": [ { ... }, ... ],    # 様式F
        "elevators": [ { ... }, ... ],     # 様式G
        "solar_pvs": [ { ... }, ... ],     # 様式H
        "cogenerations": [ { ... }, ... ], # 様式I
    }
    """
    building = input_data.get("building", {})
    # Decide by template sheet names, not floor area, so fallback template changes remain consistent.
    is_small = "様式SA_基本情報" in workbook.sheetnames

    # 様式A: 基本情報
    _write_form_a(workbook, building, is_small)

    # 様式B1〜I: テーブル形式
    table_map = {
        "B1": "windows",
        "B2": "insulations",
        "B3": "envelopes",
        "C1": "heat_sources",
        "C2": "outdoor_air",
        "C3": "pumps",
        "C4": "fans",
        "D": "ventilations",
        "E": "lightings",
        "F": "hot_waters",
        "G": "elevators",
        "H": "solar_pvs",
        "I": "cogenerations",
    }
    for form_key, data_key in table_map.items():
        rows = input_data.get(data_key, [])
        if rows:
            _write_table_rows(workbook, form_key, rows, is_small)


def _select_template(total_floor_area: float) -> Path:
    """Return the appropriate Excel template path based on floor area."""
    # NOTE:
    # As of 2026-02-09 verification against /model/1/v380, SMALLMODEL template uploads
    # return "様式A 基本情報 は必ずアップロードしてください。".
    # To keep small-area requests functional, submit MODEL template for all areas.
    if total_floor_area < 300:
        logger.warning(
            "SMALLMODEL template is currently rejected by API; using MODEL template for floor area %.2f",
            total_floor_area,
        )
    template = STANDARD_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"Excel template not found at {template}")
    return template


def _build_excel_buffer(input_data: Dict[str, Any]) -> io.BytesIO:
    """Build an in-memory Excel file from *input_data* using the official template."""
    building_data = input_data.get("building", {})
    total_area = float(
        building_data.get("calc_floor_area", 0)
        or building_data.get("total_floor_area", 0)
        or 0
    )

    template_path = _select_template(total_area)
    logger.info("Using template %s for floor area %.2f", template_path, total_area)

    workbook = openpyxl.load_workbook(template_path)
    _write_data_to_workbook(workbook, input_data)

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return buf


def _post_to_api(url: str, excel_buffer: io.BytesIO, timeout: int = 120) -> requests.Response:
    """POST an Excel buffer to a lowenergy.jp endpoint and return the response."""
    headers = {"Content-Type": EXCEL_CONTENT_TYPE}
    excel_buffer.seek(0)
    payload = excel_buffer.read()
    try:
        response = requests.post(url, data=payload, headers=headers, timeout=timeout)
        response.raise_for_status()
    except requests.exceptions.RequestException as exc:
        detail = exc.response.text if getattr(exc, "response", None) is not None else ""
        logger.exception("API call failed (%s): %s", url, detail)
        raise Exception(f"API request failed: {exc} {detail}") from exc
    return response


def _extract_api_error_message(payload: Dict[str, Any]) -> str:
    """Extract human-readable error messages from API JSON payload."""
    messages: List[str] = []

    def add_messages(items: Any) -> None:
        if not isinstance(items, list):
            return
        for item in items:
            if not isinstance(item, dict):
                continue
            msg = item.get("Message")
            if isinstance(msg, str) and msg.strip():
                messages.append(msg.strip())

    add_messages(payload.get("Errors"))
    for key, value in payload.items():
        if key.endswith("ValidationResult") and isinstance(value, dict):
            add_messages(value.get("Errors"))
            add_messages(value.get("AllInfo"))

    if messages:
        return " / ".join(dict.fromkeys(messages))
    status = payload.get("Status")
    if status and status != "OK":
        return f"Status={status}"
    return "Unknown API error"


def _extract_pdf_content_or_raise(response: requests.Response) -> bytes:
    """Return PDF bytes or raise if the API returned a JSON error payload."""
    if response.content.startswith(b"%PDF"):
        return response.content

    content_type = response.headers.get("Content-Type", "")
    if "application/json" in content_type or response.content.startswith(b"{"):
        try:
            payload = response.json()
        except ValueError:
            payload = None
        if isinstance(payload, dict):
            detail = _extract_api_error_message(payload)
            if "様式A 基本情報 は必ずアップロードしてください。" in detail:
                raise ValueError(SMALLMODEL_UPLOAD_UNSUPPORTED_MESSAGE)
            raise Exception(f"Official report API returned error: {detail}")

    raise Exception("Official report API did not return a valid PDF response.")


def _is_smallmodel_original_upload(excel_bytes: bytes) -> bool:
    """Return True when uploaded bytes look like original SMALLMODEL workbook."""
    wb: Optional[openpyxl.Workbook] = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        return "様式SA_基本情報" in wb.sheetnames and "様式A_基本情報" not in wb.sheetnames
    except Exception:
        return False
    finally:
        if wb is not None:
            wb.close()


def get_official_report_from_api(input_data: Dict[str, Any]) -> bytes:
    """Fill the official Excel template and submit to the v380 API → 公式様式PDF."""
    buf = _build_excel_buffer(input_data)
    response = _post_to_api(API_REPORT, buf)
    pdf = _extract_pdf_content_or_raise(response)
    logger.info("Received %d bytes (official PDF) from %s", len(pdf), API_REPORT)
    return pdf


def get_official_compute_from_api(input_data: Dict[str, Any]) -> Dict[str, Any]:
    """Fill the official Excel template and submit to the v380 API → 公式計算結果JSON."""
    buf = _build_excel_buffer(input_data)
    response = _post_to_api(API_COMPUTE, buf)
    logger.info("Received compute result from %s", API_COMPUTE)
    return response.json()


def get_official_report_from_excel(excel_bytes: bytes) -> bytes:
    """Submit a user-uploaded Excel input sheet directly to the v380 API → 公式様式PDF."""
    if _is_smallmodel_original_upload(excel_bytes):
        raise ValueError(SMALLMODEL_UPLOAD_UNSUPPORTED_MESSAGE)
    buf = io.BytesIO(excel_bytes)
    response = _post_to_api(API_REPORT, buf)
    pdf = _extract_pdf_content_or_raise(response)
    logger.info("Received %d bytes (official PDF from uploaded Excel)", len(pdf))
    return pdf


def get_official_compute_from_excel(excel_bytes: bytes) -> Dict[str, Any]:
    """Submit a user-uploaded Excel input sheet directly to the v380 API → 公式計算結果JSON."""
    if _is_smallmodel_original_upload(excel_bytes):
        raise ValueError(SMALLMODEL_UPLOAD_UNSUPPORTED_MESSAGE)
    buf = io.BytesIO(excel_bytes)
    response = _post_to_api(API_COMPUTE, buf)
    logger.info("Received compute result from uploaded Excel")
    return response.json()
