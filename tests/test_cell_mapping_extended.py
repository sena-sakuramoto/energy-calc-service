"""Tests for extended CELL_MAPPING coverage (B3, C1-C4, H, I)."""

import io
import unittest
from unittest.mock import MagicMock, patch

import openpyxl

from app.schemas.bei import (
    CogenerationSpec,
    EnvelopeSpec,
    FanSpec,
    HeatSourceSpec,
    OutdoorAirSpec,
    OfficialInput,
    OfficialBuildingInfo,
    PumpSpec,
    SolarPVSpec,
)
from app.services import report


class TestB3EnvelopeMapping(unittest.TestCase):
    """Test 様式B3: 外皮仕様 mapping."""

    def test_envelope_columns_defined(self):
        """Verify B3 table columns are defined."""
        self.assertIn("B3", report.TABLE_COLUMNS)
        b3_def = report.TABLE_COLUMNS["B3"]
        self.assertEqual(b3_def["sheet"], "様式B3_外皮")

    def test_envelope_all_fields_mapped(self):
        """Verify all EnvelopeSpec fields have column mappings."""
        b3_columns = report.TABLE_COLUMNS["B3"]["columns"]
        envelope_fields = set(EnvelopeSpec.model_fields.keys())
        # name is not mapped as a column in the same way
        required_fields = envelope_fields - {"name"}
        for field in required_fields:
            self.assertIn(
                field, b3_columns, f"EnvelopeSpec.{field} not mapped in B3 columns"
            )

    def test_envelope_columns_map_to_excel_columns(self):
        """Verify B3 columns map to valid Excel column letters."""
        b3_columns = report.TABLE_COLUMNS["B3"]["columns"]
        valid_columns = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
        for field, col in b3_columns.items():
            self.assertIn(col, valid_columns, f"Invalid Excel column {col} for {field}")

    def test_write_envelope_data_to_workbook(self):
        """Test writing envelope data to Excel workbook."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "envelopes": [
                {
                    "name": "Wall-South",
                    "direction": "南",
                    "width": 10.0,
                    "height": 5.0,
                    "area": 50.0,
                    "insulation_name": "断熱仕様-A",
                    "window_name": "窓-1",
                    "window_count": 5,
                    "has_blind": "有",
                    "shade_coeff_cooling": 0.6,
                    "shade_coeff_heating": 0.3,
                }
            ],
        }

        # Load a minimal MODEL template
        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式B3_外皮")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式B3_外皮"]
        # Check that data was written to the correct cells starting at row 11
        self.assertEqual(ws["A11"].value, "Wall-South")
        self.assertEqual(ws["B11"].value, "南")
        self.assertEqual(ws["C11"].value, 10.0)
        self.assertEqual(ws["D11"].value, 5.0)
        self.assertEqual(ws["E11"].value, 50.0)
        self.assertEqual(ws["F11"].value, "断熱仕様-A")
        self.assertEqual(ws["G11"].value, "窓-1")
        self.assertEqual(ws["H11"].value, 5)
        self.assertEqual(ws["I11"].value, "有")
        self.assertEqual(ws["J11"].value, 0.6)
        self.assertEqual(ws["K11"].value, 0.3)


class TestC1HeatSourceMapping(unittest.TestCase):
    """Test 様式C1: 空調熱源 mapping."""

    def test_heat_source_columns_defined(self):
        """Verify C1 table columns are defined."""
        self.assertIn("C1", report.TABLE_COLUMNS)
        c1_def = report.TABLE_COLUMNS["C1"]
        self.assertEqual(c1_def["sheet"], "様式C1_空調熱源")

    def test_heat_source_all_fields_mapped(self):
        """Verify all HeatSourceSpec fields have column mappings."""
        c1_columns = report.TABLE_COLUMNS["C1"]["columns"]
        heat_source_fields = set(HeatSourceSpec.model_fields.keys())
        for field in heat_source_fields:
            self.assertIn(
                field, c1_columns, f"HeatSourceSpec.{field} not mapped in C1 columns"
            )

    def test_write_heat_source_data_to_workbook(self):
        """Test writing heat source data to Excel workbook."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "heat_sources": [
                {
                    "name": "チラー-1",
                    "type": "ウォータチリングユニット(空冷式)",
                    "count": 2,
                    "capacity_cooling": 100.0,
                    "capacity_heating": 80.0,
                    "power_cooling": 30.0,
                    "power_heating": 25.0,
                    "fuel_cooling": 0.0,
                    "fuel_heating": 0.0,
                }
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式C1_空調熱源")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式C1_空調熱源"]
        self.assertEqual(ws["A11"].value, "チラー-1")
        self.assertEqual(ws["B11"].value, "ウォータチリングユニット(空冷式)")
        self.assertEqual(ws["C11"].value, 2)
        self.assertEqual(ws["D11"].value, 100.0)
        self.assertEqual(ws["E11"].value, 80.0)
        self.assertEqual(ws["F11"].value, 30.0)
        self.assertEqual(ws["G11"].value, 25.0)
        self.assertEqual(ws["H11"].value, 0.0)
        self.assertEqual(ws["I11"].value, 0.0)


class TestC2OutdoorAirMapping(unittest.TestCase):
    """Test 様式C2: 空調外気処理 mapping."""

    def test_outdoor_air_columns_defined(self):
        """Verify C2 table columns are defined."""
        self.assertIn("C2", report.TABLE_COLUMNS)
        c2_def = report.TABLE_COLUMNS["C2"]
        self.assertEqual(c2_def["sheet"], "様式C2_空調外気処理")

    def test_outdoor_air_all_fields_mapped(self):
        """Verify all OutdoorAirSpec fields have column mappings."""
        c2_columns = report.TABLE_COLUMNS["C2"]["columns"]
        outdoor_air_fields = set(OutdoorAirSpec.model_fields.keys())
        for field in outdoor_air_fields:
            self.assertIn(
                field, c2_columns, f"OutdoorAirSpec.{field} not mapped in C2 columns"
            )

    def test_write_outdoor_air_data_to_workbook(self):
        """Test writing outdoor air processing data to Excel workbook."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "outdoor_air": [
                {
                    "name": "外気処理-1",
                    "count": 1,
                    "supply_airflow": 2000.0,
                    "exhaust_airflow": 1800.0,
                    "heat_exchange_eff_cooling": 65.0,
                    "heat_exchange_eff_heating": 70.0,
                    "auto_bypass": "有",
                    "preheat_stop": "無",
                }
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式C2_空調外気処理")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式C2_空調外気処理"]
        self.assertEqual(ws["A11"].value, "外気処理-1")
        self.assertEqual(ws["B11"].value, 1)
        self.assertEqual(ws["C11"].value, 2000.0)
        self.assertEqual(ws["D11"].value, 1800.0)
        self.assertEqual(ws["E11"].value, 65.0)
        self.assertEqual(ws["F11"].value, 70.0)
        self.assertEqual(ws["G11"].value, "有")
        self.assertEqual(ws["H11"].value, "無")


class TestC3PumpMapping(unittest.TestCase):
    """Test 様式C3: 空調二次ポンプ mapping (MODEL固有)."""

    def test_pump_columns_defined(self):
        """Verify C3 table columns are defined."""
        self.assertIn("C3", report.TABLE_COLUMNS)
        c3_def = report.TABLE_COLUMNS["C3"]
        self.assertEqual(c3_def["sheet"], "様式C3_空調ポンプ")

    def test_pump_all_fields_mapped(self):
        """Verify all PumpSpec fields have column mappings."""
        c3_columns = report.TABLE_COLUMNS["C3"]["columns"]
        pump_fields = set(PumpSpec.model_fields.keys())
        for field in pump_fields:
            self.assertIn(
                field, c3_columns, f"PumpSpec.{field} not mapped in C3 columns"
            )

    def test_pump_skipped_for_smallmodel(self):
        """Verify C3 is skipped when writing to SMALLMODEL template."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 200,  # < 300, would use SMALLMODEL
            },
            "pumps": [
                {
                    "name": "ポンプ-1",
                    "count": 1,
                    "flow_rate": 50.0,
                    "variable_flow": "有",
                    "min_flow_input": "入力する",
                    "min_flow_ratio": 30.0,
                }
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式SA_基本情報"  # SMALLMODEL sheet name
        # Note: SMALLMODEL doesn't have C3, so _write_data_to_workbook should skip it

        report._write_data_to_workbook(wb, input_data)
        # Should not raise an error, just skip the C3 form


class TestC4FanMapping(unittest.TestCase):
    """Test 様式C4: 空調送風機 mapping (MODEL固有)."""

    def test_fan_columns_defined(self):
        """Verify C4 table columns are defined."""
        self.assertIn("C4", report.TABLE_COLUMNS)
        c4_def = report.TABLE_COLUMNS["C4"]
        self.assertEqual(c4_def["sheet"], "様式C4_空調送風機")

    def test_fan_all_fields_mapped(self):
        """Verify all FanSpec fields have column mappings."""
        c4_columns = report.TABLE_COLUMNS["C4"]["columns"]
        fan_fields = set(FanSpec.model_fields.keys())
        for field in fan_fields:
            self.assertIn(
                field, c4_columns, f"FanSpec.{field} not mapped in C4 columns"
            )

    def test_write_fan_data_to_workbook(self):
        """Test writing fan data to Excel workbook."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "fans": [
                {
                    "name": "送風機-1",
                    "count": 2,
                    "airflow": 1000.0,
                    "variable_airflow": "有",
                    "min_airflow_input": "入力する",
                    "min_airflow_ratio": 40.0,
                }
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式C4_空調送風機")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式C4_空調送風機"]
        self.assertEqual(ws["A11"].value, "送風機-1")
        self.assertEqual(ws["B11"].value, 2)
        self.assertEqual(ws["C11"].value, 1000.0)
        self.assertEqual(ws["D11"].value, "有")
        self.assertEqual(ws["E11"].value, "入力する")
        self.assertEqual(ws["F11"].value, 40.0)


class TestHSolarPVMapping(unittest.TestCase):
    """Test 様式H: 太陽光発電 mapping."""

    def test_solar_pv_columns_defined(self):
        """Verify H table columns are defined."""
        self.assertIn("H", report.TABLE_COLUMNS)
        h_def = report.TABLE_COLUMNS["H"]
        self.assertEqual(h_def["sheet"], "様式H_太陽光発電")

    def test_solar_pv_all_fields_mapped(self):
        """Verify all SolarPVSpec fields have column mappings."""
        h_columns = report.TABLE_COLUMNS["H"]["columns"]
        solar_pv_fields = set(SolarPVSpec.model_fields.keys())
        for field in solar_pv_fields:
            self.assertIn(
                field, h_columns, f"SolarPVSpec.{field} not mapped in H columns"
            )

    def test_write_solar_pv_data_to_workbook(self):
        """Test writing solar PV data to Excel workbook."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "solar_pvs": [
                {
                    "system_name": "太陽光-1",
                    "cell_type": "結晶系太陽電池",
                    "installation_mode": "屋根置き形",
                    "capacity_kw": 10.0,
                    "panel_direction": "0度(南)",
                    "panel_angle": "30度",
                }
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式H_太陽光発電")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式H_太陽光発電"]
        self.assertEqual(ws["A11"].value, "太陽光-1")
        self.assertEqual(ws["B11"].value, "結晶系太陽電池")
        self.assertEqual(ws["C11"].value, "屋根置き形")
        self.assertEqual(ws["D11"].value, 10.0)
        self.assertEqual(ws["E11"].value, "0度(南)")
        self.assertEqual(ws["F11"].value, "30度")


class TestICogenerationMapping(unittest.TestCase):
    """Test 様式I: コージェネレーション設備 mapping (MODEL固有)."""

    def test_cogeneration_columns_defined(self):
        """Verify I table columns are defined."""
        self.assertIn("I", report.TABLE_COLUMNS)
        i_def = report.TABLE_COLUMNS["I"]
        self.assertEqual(i_def["sheet"], "様式I_コージェネレーション設備")

    def test_cogeneration_all_fields_mapped(self):
        """Verify all CogenerationSpec fields have column mappings."""
        i_columns = report.TABLE_COLUMNS["I"]["columns"]
        cogeneration_fields = set(CogenerationSpec.model_fields.keys())
        for field in cogeneration_fields:
            self.assertIn(
                field, i_columns, f"CogenerationSpec.{field} not mapped in I columns"
            )

    def test_write_cogeneration_data_to_workbook(self):
        """Test writing cogeneration data to Excel workbook."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "cogenerations": [
                {
                    "name": "コージェネ-1",
                    "rated_output": 50.0,
                    "count": 1,
                    "gen_eff_100": 32.0,
                    "gen_eff_75": 30.0,
                    "gen_eff_50": 28.0,
                    "heat_eff_100": 42.0,
                    "heat_eff_75": 44.0,
                    "heat_eff_50": 45.0,
                    "heat_recovery_for": "冷房と暖房と給湯",
                }
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式I_コージェネレーション設備")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式I_コージェネレーション設備"]
        self.assertEqual(ws["A11"].value, "コージェネ-1")
        self.assertEqual(ws["B11"].value, 50.0)
        self.assertEqual(ws["C11"].value, 1)
        self.assertEqual(ws["D11"].value, 32.0)
        self.assertEqual(ws["E11"].value, 30.0)
        self.assertEqual(ws["F11"].value, 28.0)
        self.assertEqual(ws["G11"].value, 42.0)
        self.assertEqual(ws["H11"].value, 44.0)
        self.assertEqual(ws["I11"].value, 45.0)
        self.assertEqual(ws["J11"].value, "冷房と暖房と給湯")


class TestMultipleTableRows(unittest.TestCase):
    """Test writing multiple rows to table forms."""

    def test_write_multiple_heat_sources(self):
        """Test writing multiple heat source rows."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "heat_sources": [
                {
                    "name": "チラー-1",
                    "type": "ウォータチリングユニット(空冷式)",
                    "count": 2,
                    "capacity_cooling": 100.0,
                    "capacity_heating": 80.0,
                    "power_cooling": 30.0,
                    "power_heating": 25.0,
                    "fuel_cooling": 0.0,
                    "fuel_heating": 0.0,
                },
                {
                    "name": "ボイラ-1",
                    "type": "ボイラ",
                    "count": 1,
                    "capacity_cooling": 0.0,
                    "capacity_heating": 150.0,
                    "power_cooling": 0.0,
                    "power_heating": 5.0,
                    "fuel_cooling": 0.0,
                    "fuel_heating": 100.0,
                },
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式C1_空調熱源")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式C1_空調熱源"]
        # First row (row 11)
        self.assertEqual(ws["A11"].value, "チラー-1")
        self.assertEqual(ws["B11"].value, "ウォータチリングユニット(空冷式)")
        # Second row (row 12)
        self.assertEqual(ws["A12"].value, "ボイラ-1")
        self.assertEqual(ws["B12"].value, "ボイラ")

    def test_write_multiple_solar_pv_systems(self):
        """Test writing multiple solar PV rows."""
        input_data = {
            "building": {
                "building_name": "Test Building",
                "region": "6地域",
                "building_type": "事務所モデル",
                "calc_floor_area": 500,
            },
            "solar_pvs": [
                {
                    "system_name": "太陽光-1",
                    "cell_type": "結晶系太陽電池",
                    "installation_mode": "屋根置き形",
                    "capacity_kw": 10.0,
                    "panel_direction": "0度(南)",
                    "panel_angle": "30度",
                },
                {
                    "system_name": "太陽光-2",
                    "cell_type": "結晶系太陽電池",
                    "installation_mode": "架台設置形",
                    "capacity_kw": 5.0,
                    "panel_direction": "90度(西)",
                    "panel_angle": "30度",
                },
            ],
        }

        wb = openpyxl.Workbook()
        wb.active.title = "様式A_基本情報"
        wb.create_sheet("様式H_太陽光発電")

        report._write_data_to_workbook(wb, input_data)

        ws = wb["様式H_太陽光発電"]
        # First row
        self.assertEqual(ws["A11"].value, "太陽光-1")
        self.assertEqual(ws["D11"].value, 10.0)
        # Second row
        self.assertEqual(ws["A12"].value, "太陽光-2")
        self.assertEqual(ws["D12"].value, 5.0)


class TestSmallModelSheetMapping(unittest.TestCase):
    """Test SMALLMODEL vs MODEL sheet name mapping."""

    def test_b3_skipped_in_smallmodel(self):
        """Verify B3 is None in SMALL_SHEET_MAP (skipped for SMALLMODEL)."""
        self.assertIsNone(report.SMALL_SHEET_MAP.get("様式B3_外皮"))

    def test_c3_skipped_in_smallmodel(self):
        """Verify C3 is None in SMALL_SHEET_MAP (skipped for SMALLMODEL)."""
        self.assertIsNone(report.SMALL_SHEET_MAP.get("様式C3_空調ポンプ"))

    def test_c4_skipped_in_smallmodel(self):
        """Verify C4 is None in SMALL_SHEET_MAP (skipped for SMALLMODEL)."""
        self.assertIsNone(report.SMALL_SHEET_MAP.get("様式C4_空調送風機"))

    def test_g_skipped_in_smallmodel(self):
        """Verify G is None in SMALL_SHEET_MAP (skipped for SMALLMODEL)."""
        self.assertIsNone(report.SMALL_SHEET_MAP.get("様式G_昇降機"))

    def test_i_skipped_in_smallmodel(self):
        """Verify I is None in SMALL_SHEET_MAP (skipped for SMALLMODEL)."""
        self.assertIsNone(report.SMALL_SHEET_MAP.get("様式I_コージェネレーション設備"))

    def test_h_available_in_smallmodel(self):
        """Verify H (solar PV) is available in SMALLMODEL."""
        self.assertEqual(
            report.SMALL_SHEET_MAP.get("様式H_太陽光発電"), "様式SH_太陽光発電"
        )


if __name__ == "__main__":
    unittest.main()
